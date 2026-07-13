#!/usr/bin/env python3
"""Export the HSK level reference to a PORTABLE snapshot the backend can load anywhere.

WHY: the pipeline's own level checker (Scripts/Research and Writing/HSK Mandarin/stages/
level_check.py) reads hsk_vocab.xlsx / proper_nouns.xlsx — and `*.xlsx` is GITIGNORED in
the dynamic-content repo, so those files exist ONLY on the machine that built them. The
live review host (the laptop) has the Scripts checkout but NOT the xlsx, so the backend
cannot import level_check there. Instead we snapshot the reference to JSON, commit it to
THIS repo, and let backend/app/zh_level.py read that (jieba is in the live venv; openpyxl
and the xlsx are not needed at runtime).

Run on a machine that HAS the xlsx (the workstation), then commit the output:
    py -3.12 scripts/export_hsk_vocab.py
    -> backend/app/data/hsk_vocab.json

Re-run whenever the HSK reference or hsk_config's proper-noun lists change.
"""
from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "backend" / "app" / "data" / "hsk_vocab.json"

SCRIPTS_ROOT = Path(os.environ.get("REVIEW_APP_SCRIPTS_ROOT", r"D:\Dynamic Languages\Scripts"))
HSK_DIR = SCRIPTS_ROOT / "Research and Writing" / "HSK Mandarin"

# hsk_config asserts a loc+gdoc for the CURRENT trip; pin dummies so importing it for its
# TRIPS / _BASE_PROPER tables never depends on which trip happens to be selected.
os.environ.setdefault("HSK_LOC", "x")
os.environ.setdefault("HSK_GDOC", "x")
sys.path.insert(0, str(HSK_DIR))
sys.path.insert(0, str(HSK_DIR / "stages"))

import openpyxl  # noqa: E402
import hsk_config as cfg  # noqa: E402


def load_surfaces() -> dict[str, str]:
    """simplified/traditional surface -> HSK level label (mirrors level_check.load_vocab)."""
    surfaces: dict[str, str] = {}
    wb = openpyxl.load_workbook(cfg.VOCAB_XLSX, read_only=True)
    ws = wb.active
    for simp, trad, pinyin, meaning, pos, level in ws.iter_rows(min_row=2, values_only=True):
        for field in (simp, trad):
            if not field:
                continue
            for variant in str(field).split("|"):      # 爸爸|爸 -> two surfaces
                v = variant.strip()
                if v:
                    surfaces.setdefault(v, level)
    wb.close()
    if os.path.exists(cfg.VOCAB_SUPPLEMENT_TSV):
        for ln in open(cfg.VOCAB_SUPPLEMENT_TSV, encoding="utf-8"):
            ln = ln.strip()
            if not ln or ln.startswith("#"):
                continue
            parts = ln.split("\t")
            if len(parts) >= 3:
                surfaces.setdefault(parts[0].strip(), parts[2].strip())
    return surfaces


def load_proper_base() -> list[str]:
    names = set(cfg._BASE_PROPER)
    if os.path.exists(cfg.PROPER_NOUNS_XLSX):
        wb = openpyxl.load_workbook(cfg.PROPER_NOUNS_XLSX, read_only=True)
        ws = wb.active
        for surface, *_ in ws.iter_rows(min_row=2, values_only=True):
            if surface:
                names.add(str(surface).strip())
        wb.close()
    return sorted(names)


def main() -> None:
    surfaces = load_surfaces()
    # Per-trip proper nouns stay PER TRIP (not unioned): _is_proper does substring
    # matching, so pooling every trip's names would exempt one trip's word in another.
    proper_by_trip = {k: sorted(v.get("proper") or []) for k, v in cfg.TRIPS.items()}
    snap = {
        "generated_at": time.time(),
        "source": "hsk_vocab.xlsx + vocab_supplement.tsv + proper_nouns.xlsx (dynamic-content)",
        "bands": {k: sorted(v) for k, v in cfg.IN_BAND_BY_LEVEL.items()},
        "surfaces": surfaces,
        "proper_base": load_proper_base(),
        "proper_by_trip": proper_by_trip,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(snap, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"wrote {OUT}  ({len(surfaces)} surfaces, {len(snap['proper_base'])} base proper "
          f"nouns, {len(proper_by_trip)} trips, bands={list(snap['bands'])})")


if __name__ == "__main__":
    main()
